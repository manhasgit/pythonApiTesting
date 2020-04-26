import requests
import json
import jsonpath
import openpyxl

def test_add_student_multiple():
    global request_json
    API_URL = 'http://www.thetestingworldapi.com/api/studentsDetails'
    file = open('//Users//nareshmanhas//Desktop//hello//API//Add_Request.json','r')
    path = '//Users//nareshmanhas//Desktop//hello//API//Api_Excel.xlsx'
    workbook = openpyxl.load_workbook(path)
    sheet = workbook['Sheet1']
    rows = sheet.max_row
    request_json = json.loads(file.read())
    for i in range(2, rows + 1):
        cell_first_name = sheet.cell(row=i,column=1)
        cell_mid_name = sheet.cell(row=i,column=2)
        cell_last_name = sheet.cell(row=i,column=3)
        cell_dob_name = sheet.cell(row=i,column=4)
        request_json['first_name'] = cell_first_name.value
        request_json['middle_name'] = cell_mid_name.value
        request_json['last_name'] = cell_last_name.value
        request_json['date_of_birth'] = cell_dob_name.value

        response = requests.post(API_URL, request_json)
        print(response.text)
        print(response.status_code)
        assert response.status_code == 201
